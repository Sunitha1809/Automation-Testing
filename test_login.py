import pytest
import allure
from openpyxl import load_workbook
from Pages.login_page import LoginPage


# 📌 Read Excel Data
def get_test_data():
    workbook = load_workbook("Fixtures/login_data.xlsx")
    sheet = workbook.active

    data = []
    rows = sheet.max_row

    for i in range(2, rows + 1):
        username = sheet.cell(row=i, column=1).value
        password = sheet.cell(row=i, column=2).value
        expected = sheet.cell(row=i, column=3).value
        data.append((username, password, expected))

    return data


@pytest.mark.parametrize("username,password,expected", get_test_data())
@allure.title("SauceDemo Login Data Driven Test")
def test_login(driver, username, password, expected):

    driver.get("https://www.saucedemo.com")

    login_page = LoginPage(driver)
    login_page.login(username, password)

    if expected:
        assert "inventory" in driver.current_url
    else:
        assert "inventory" not in driver.current_url
